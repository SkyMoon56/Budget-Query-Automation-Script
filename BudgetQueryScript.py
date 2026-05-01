# ============================================================
# Budget Query Automation
# ============================================================
# PURPOSE:
#   Combines the OU_BUD_ORG and OU_BUD_SOURCE PeopleSoft query
#   exports into a formatted Excel file each month.
#
# HOW TO RUN:
#   1. Export OU_BUD_ORG and OU_BUD_SOURCE from PeopleSoft
#   2. Run this script (py budget_query_automation.py)
#   3. A file picker will open -- select OU_BUD_ORG first
#   4. A second file picker will open -- select OU_BUD_SOURCE
#   5. Enter the department name when prompted (ex: Biology)
#   6. The finished Excel file saves in the same folder as this script
#      named: P## Month Year - Department.xlsx
#
# REQUIREMENTS:
#   py -m pip install pandas openpyxl
# ============================================================

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import tkinter as tk
from tkinter import filedialog, simpledialog


def get_period_and_month():
    """
    Determines the correct period number, month name, and year for the filename.
    Budget queries are always named for the PREVIOUS month.
    Example: Running on April 3 -> P03 March 2026
    """
    today = date.today()
    first_of_this_month = today.replace(day=1)
    last_month = first_of_this_month - timedelta(days=1)
    period = last_month.month
    period_str = f"P{period:02d}"
    month_name = last_month.strftime("%B")
    year = last_month.year
    return period_str, month_name, year


def fill_business_unit(df):
    """
    Finds the first non-blank value in the Unit column and fills it across
    all rows in the combined dataframe. Called after concat so blanks from
    either file get covered by the known BU value from the other.

    Also drops the redundant BusUnit column PeopleSoft appends at the end.
    """
    if "Unit" in df.columns:
        non_blank = df["Unit"].dropna()
        non_blank = non_blank[non_blank.str.strip() != ""]
        if not non_blank.empty:
            df["Unit"] = non_blank.iloc[0]

    df = df.drop(columns=["BusUnit"], errors="ignore")
    return df


def load_and_sort_org(path):
    """
    Loads the OU_BUD_ORG export. PeopleSoft puts a title row at row 0
    and the real headers at row 1, so we skip row 0 and use row 1 as headers.

    Custom sort applied:
      1. Fund (A-Z)
      2. Function (A-Z)
      3. Budget Type (Z-A)
      4. Account (A-Z)
    """
    df = pd.read_excel(path, dtype=str, header=1)
    df = df.sort_values(
        by=["Fund", "Function", "Budget Type", "Account"],
        ascending=[True, True, False, True]
    )
    return df


def load_and_sort_source(path):
    """
    Loads the OU_BUD_SOURCE export. Same PeopleSoft header structure as ORG.
    Removes OUFND rows to avoid duplicating data already in OU_BUD_ORG.

    Custom sort applied:
      1. Fund (A-Z)
      2. Source (A-Z)
      3. Function (A-Z)
      4. Budget Type (Z-A)
      5. Account (A-Z)
    """
    df = pd.read_excel(path, dtype=str, header=1)
    df = df[df["Fund"] != "OUFND"]
    df = df.sort_values(
        by=["Fund", "Source", "Function", "Budget Type", "Account"],
        ascending=[True, True, True, False, True]
    )
    return df


def write_workbook(df, department, period_str, month_name, year):
    """
    Writes the combined dataframe to a formatted Excel file.

    Formatting applied:
      - A1: "Org Budget Inquiry" label to match reference format
      - B1: "Retrieved MM/DD/YYYY" -- bold, yellow highlight
      - Row 2: Column headers -- bold, with auto-filter
      - Dollar columns: Accounting format ($ separated, values right-aligned)
      - Parent rows: Light green highlight across all columns
      - Column widths: Auto-fit to content (max 50)
    """
    filename = f"{period_str} {month_name} {year} - {department}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sheet1"

    ws["A1"] = "Org Budget Inquiry"

    today_str = date.today().strftime("%m/%d/%Y")
    ws["B1"] = f"Retrieved {today_str}"
    ws["B1"].font = Font(bold=True)
    ws["B1"].fill = PatternFill("solid", start_color="FFFF00")

    headers = list(df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)

    usd_cols = []
    usd_keywords = ["budget", "amount", "remaining", "actual", "encumb"]
    for col_idx, header in enumerate(headers, start=1):
        if any(k in header.lower() for k in usd_keywords):
            usd_cols.append(col_idx)

    green_fill = PatternFill("solid", start_color="C6EFCE")

    for row_idx, row in enumerate(df.itertuples(index=False), start=3):
        is_parent_row = False

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if col_idx in usd_cols:
                try:
                    cell.value = float(str(value).replace(",", "").replace("$", ""))
                    cell.number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
                except (ValueError, TypeError):
                    pass

            if str(value).strip().upper() == "PARENT":
                is_parent_row = True

        if is_parent_row:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=c).fill = green_fill

    ws.auto_filter.ref = (
        ws.cell(row=2, column=1).coordinate
        + ":"
        + ws.cell(row=2, column=len(headers)).coordinate
    )

    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    wb.save(filename)
    print(f"Saved: {filename}")


def pick_file(title):
    """Opens a file picker dialog and returns the selected file path."""
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askopenfilename(title=title, filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv")])
    root.destroy()
    if not path:
        raise SystemExit(f"No file selected for: {title}")
    return path


def prompt_department():
    """Prompts the user to type in the department name for the filename."""
    root = tk.Tk()
    root.withdraw()
    name = simpledialog.askstring("Department", "Enter department name (ex: Biology):")
    root.destroy()
    if not name:
        raise SystemExit("No department name entered.")
    return name.strip()


def main():
    period_str, month_name, year = get_period_and_month()

    org_path = pick_file("Select OU_BUD_ORG file")
    source_path = pick_file("Select OU_BUD_SOURCE file")

    department = prompt_department()

    org_df = load_and_sort_org(org_path)
    source_df = load_and_sort_source(source_path)
    combined = pd.concat([org_df, source_df], ignore_index=True)

    # Fill Business Unit and drop BusUnit AFTER concat so blanks from
    # either file get covered by the known BU value from the other
    combined = fill_business_unit(combined)

    write_workbook(combined, department, period_str, month_name, year)


if __name__ == "__main__":
    main()
